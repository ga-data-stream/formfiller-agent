from contextlib import contextmanager

from formfiller.agent.models import LoopOutcome
from formfiller.agent.pipeline import run_agent_pipeline, AgentDeps
from formfiller.config import AppConfig, ProfileField
from formfiller.models import EmailMessage, FormSchema, MappingResult, MappingOutcome
from formfiller.orchestrator import PipelineHooks


def _email(body="link https://forms.office.com/r/x"):
    return EmailMessage(entry_id="E1", sender="c@acme.com", subject="s",
                        received="2026-06-10T09:00:00", body_text=body, body_html="")


def _config(tmp_path, dry_run=True):
    return AppConfig(excel_log_path=str(tmp_path / "log.xlsx"),
                     review_queue_dir=str(tmp_path / "queue"),
                     traces_dir=str(tmp_path / "traces"),
                     decisions_dir=str(tmp_path / "decisions"),
                     dry_run=dry_run, fill_strategy="agent",
                     azure_openai_deployment="gpt-5.4-nano")


_PROFILE = (ProfileField(name="siren", value="123456789", aliases=()),)


def _deps(outcome):
    @contextmanager
    def open_page():
        yield object()
    return AgentDeps(
        open_page=open_page,
        run=lambda page, url, config, profile, trace: outcome,
    )


def test_dry_run_outcome_logs_success_and_saves_preview(tmp_path):
    out = LoopOutcome(status="dry_run", reason="dry-run", fields_filled=2, steps=4,
                      screenshot=b"\x89PNG")
    result = run_agent_pipeline(_email(), _config(tmp_path), _PROFILE, _deps(out),
                                det_hooks=None)
    assert result.status == "success"
    assert result.fields_filled == 2
    preview = tmp_path / "dry_run_preview" / "E1.png"
    assert preview.exists() and preview.read_bytes() == b"\x89PNG"


def test_dry_run_outcome_with_zero_fields_is_not_success(tmp_path):
    # Same invariant as the deterministic pipeline: a dry-run/submitted outcome
    # that filled nothing must not be reported as success.
    out = LoopOutcome(status="dry_run", reason="dry-run", fields_filled=0, steps=4,
                      screenshot=b"\x89PNG")
    result = run_agent_pipeline(_email(), _config(tmp_path), _PROFILE, _deps(out),
                                det_hooks=None)
    assert result.status == "fail"
    assert result.fields_filled == 0


def test_review_outcome_parks_and_logs_manual(tmp_path):
    schema = FormSchema(url="https://forms.office.com/r/x", title="t", questions=())
    out = LoopOutcome(status="review", reason="captcha", steps=1,
                      screenshot=b"\x89PNG", schema=schema, mapping=MappingResult(answers=()))
    result = run_agent_pipeline(_email(), _config(tmp_path), _PROFILE, _deps(out),
                                det_hooks=None)
    assert result.status == "manual"
    assert "captcha" in result.review_reason
    assert (tmp_path / "queue" / "E1").exists()


def test_no_link_logs_fail(tmp_path):
    out = LoopOutcome(status="dry_run", reason="x")
    result = run_agent_pipeline(_email("no link"), _config(tmp_path), _PROFILE, _deps(out),
                                det_hooks=None)
    assert result.status == "fail"
    assert "link" in result.review_reason.lower()


def test_review_without_screenshot_logs_empty_path(tmp_path):
    out = LoopOutcome(status="review", reason="login wall", steps=1)  # no screenshot
    result = run_agent_pipeline(_email(), _config(tmp_path), _PROFILE, _deps(out),
                                det_hooks=None)
    assert result.status == "manual"
    assert result.screenshot_path == ""   # no file was written, so path must be empty
    # and no screenshot.png exists in the queue dir
    assert not (tmp_path / "queue" / "E1" / "screenshot.png").exists()


def test_abort_falls_back_to_deterministic(tmp_path):
    out = LoopOutcome(status="abort", reason="max steps")
    from formfiller.models import FormQuestion, QuestionType, MappedAnswer
    schema = FormSchema(url="https://forms.office.com/r/x", title="t",
                        questions=(FormQuestion(id="q1", label="SIREN",
                                   type=QuestionType.TEXT, required=True),))
    mapping = MappingResult(answers=(MappedAnswer(question_id="q1", profile_field="siren",
              value="123456789", confidence=0.95, status="matched"),))
    det_hooks = PipelineHooks(read_form=lambda url: schema,
                              map_fields=lambda s: MappingOutcome(result=mapping, decisions=()),
                              fill_and_submit=lambda url, instr, dry: (b"\x89PNG", False, len(instr)))
    result = run_agent_pipeline(_email(), _config(tmp_path, dry_run=True), _PROFILE,
                                _deps(out), det_hooks=det_hooks)
    assert result.status == "success"
    from openpyxl import load_workbook
    rows = load_workbook(tmp_path / "log.xlsx").active.max_row
    assert rows == 2  # header + exactly one data row (the fallback's), not two
