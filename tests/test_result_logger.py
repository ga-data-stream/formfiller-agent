from pathlib import Path

import pytest
from openpyxl import load_workbook
from formfiller import result_logger
from formfiller.result_logger import JobResult, append_result, COLUMNS


def _result(status="success"):
    return JobResult(
        timestamp="2026-06-10T09:30:00",
        sender="client@acme.com",
        client_name="Acme",
        form_url="https://forms.office.com/r/x",
        form_type="ms_forms",
        status=status,
        overall_confidence=0.91,
        fields_filled=5,
        fields_blank_flagged="nickname",
        review_reason="",
        screenshot_path="",
    )


def test_creates_workbook_with_header_when_missing(tmp_path):
    path = tmp_path / "log.xlsx"
    append_result(path, _result())
    wb = load_workbook(path)
    ws = wb.active
    assert [c.value for c in ws[1]] == list(COLUMNS)
    assert ws[2][0].value == "2026-06-10T09:30:00"
    assert ws[2][5].value == "success"


def test_appends_second_row_without_duplicating_header(tmp_path):
    path = tmp_path / "log.xlsx"
    append_result(path, _result(status="success"))
    append_result(path, _result(status="manual"))
    wb = load_workbook(path)
    ws = wb.active
    assert ws.max_row == 3  # header + 2 rows
    assert ws[3][5].value == "manual"


def test_append_result_returns_path_written_on_success(tmp_path):
    path = tmp_path / "log.xlsx"
    used = result_logger.append_result(path, _result())
    assert used == path


def test_append_result_falls_back_to_sidecar_when_primary_locked(tmp_path, monkeypatch):
    primary = tmp_path / "log.xlsx"
    real_append_row = result_logger._append_row

    def fake_append_row(p, r):
        if Path(p) == primary:
            raise PermissionError("file is open in Excel")
        return real_append_row(p, r)

    monkeypatch.setattr(result_logger, "_append_row", fake_append_row)
    used = result_logger.append_result(primary, _result(), retries=1, delay=0)
    assert used != primary           # wrote a sidecar instead
    assert used.exists()
    assert used.suffix == ".xlsx"
    assert not primary.exists()      # primary was never written (it was "locked")
