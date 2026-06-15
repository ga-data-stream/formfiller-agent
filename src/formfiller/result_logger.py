from __future__ import annotations

import time
from pathlib import Path

from openpyxl import Workbook, load_workbook
from pydantic import BaseModel, ConfigDict

COLUMNS = (
    "timestamp",
    "sender",
    "client_name",
    "form_url",
    "form_type",
    "status",
    "overall_confidence",
    "fields_filled",
    "fields_blank_flagged",
    "review_reason",
    "screenshot_path",
)


class JobResult(BaseModel):
    model_config = ConfigDict(frozen=True)
    timestamp: str
    sender: str
    client_name: str
    form_url: str
    form_type: str
    status: str  # "success" | "manual" | "fail"
    overall_confidence: float
    fields_filled: int
    fields_blank_flagged: str
    review_reason: str
    screenshot_path: str


def _append_row(path, result) -> None:
    path = Path(path)
    if path.exists():
        wb = load_workbook(path)
        ws = wb.active
    else:
        path.parent.mkdir(parents=True, exist_ok=True)
        wb = Workbook()
        ws = wb.active
        ws.title = "log"
        ws.append(list(COLUMNS))
    ws.append([getattr(result, col) for col in COLUMNS])
    wb.save(path)


def _sidecar_path(path: Path, result) -> Path:
    # build a filesystem-safe stamp from the result timestamp (ISO8601)
    stamp = "".join(c for c in result.timestamp if c.isalnum()) or "log"
    return path.with_name(f"{path.stem}.locked-{stamp}{path.suffix}")


def append_result(path, result, retries: int = 3, delay: float = 0.5) -> Path:
    """Append one outcome row to the Excel log. Resilient to the file being
    locked (open in Excel / OneDrive sync): retries briefly, then writes a
    timestamped sidecar file alongside instead of failing. Returns the path
    actually written."""
    target = Path(path)
    for attempt in range(retries):
        try:
            _append_row(target, result)
            return target
        except PermissionError:
            if attempt < retries - 1 and delay:
                time.sleep(delay)
    fallback = _sidecar_path(target, result)
    _append_row(fallback, result)
    return fallback
